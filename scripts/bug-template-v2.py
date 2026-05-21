"""
Aggiorna il BUG_EXCEL_TEMPLATE.xlsx alla v2 inserendo la colonna `origine`
(enum tecnico|funzionale) usata da sdlc-debug nella Fase 2c (due ondate:
test tecnici + test funzionali manuali).

Sorgente: Downloads/New Way of working/BUG_EXCEL_TEMPLATE.xlsx
Target:   claude-flow/templates/BUG_EXCEL_TEMPLATE.xlsx

Modifiche:
  - Foglio Bugs:
      * Inserisce colonna `origine` in posizione J (dopo `tipo`).
      * Aggiunge data validation enum 'tecnico,funzionale' su tutto il range.
      * Marca le 3 righe esempio esistenti come origine=funzionale.
      * Aggiunge 1 nuova riga esempio (id=4) con origine=tecnico (perf regression).
  - Foglio Istruzioni:
      * Aggiunge bullet sulla colonna `origine` allo schema.
      * Aggiorna riferimento da `Fase 11` a `Fase 2c (test e chiusura)`.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

SRC = Path(r"C:\Users\davmelis\Downloads\New Way of working\BUG_EXCEL_TEMPLATE.xlsx")
DST = Path(r"C:\Users\davmelis\Documents\MyGitHub\claude-flow\templates\BUG_EXCEL_TEMPLATE.xlsx")

wb = load_workbook(SRC)

# --- Sheet Bugs ---
ws = wb["Bugs"]

# 1. Inserisci colonna J (10) — 'origine' dopo 'tipo' (col I=9)
ws.insert_cols(idx=10)
hdr_cell = ws.cell(row=1, column=10, value="origine")
hdr_cell.font = Font(bold=True)

# 2. Popola origine sulle 3 righe esempio esistenti come 'funzionale'
for r in range(2, ws.max_row + 1):
    if ws.cell(row=r, column=1).value is not None:  # ha id
        ws.cell(row=r, column=10, value="funzionale")

# 3. Aggiunge 1 esempio tecnico (id 4)
new_row_idx = ws.max_row + 1
ws.cell(row=new_row_idx, column=1, value=4)                           # id
ws.cell(row=new_row_idx, column=2, value="Prenotazione")              # fase
ws.cell(row=new_row_idx, column=3, value="API booking")               # sezione
ws.cell(row=new_row_idx, column=4, value="-")                          # utente (n/a per tecnico)
ws.cell(row=new_row_idx, column=5,
        value="POST /api/bookings p95 latency > 2s sotto carico 500 rps")
ws.cell(row=new_row_idx, column=6,
        value=("Performance regression: load test k6 con 500 rps rivela p95=2.3s "
               "(budget=500ms). Profilo: query N+1 su availability check."))
ws.cell(row=new_row_idx, column=7, value="reports/k6-2026-05-20.html")  # screenshot
ws.cell(row=new_row_idx, column=8, value=None)                         # riferimento
ws.cell(row=new_row_idx, column=9, value="DEFECT/BUG")                 # tipo
ws.cell(row=new_row_idx, column=10, value="tecnico")                   # origine (nuova)
ws.cell(row=new_row_idx, column=11, value="Aperto")                    # stato_originale
ws.cell(row=new_row_idx, column=12, value="2026-05-20")                # data
ws.cell(row=new_row_idx, column=13, value=None)                        # note_dev
ws.cell(row=new_row_idx, column=14,
        value="Rilevato in run k6 ondata (a) test tecnici. Bug non visibile a UI.")

# 4. Data validation enum su intera colonna J (2..1000 per coprire usi futuri)
dv = DataValidation(type="list", formula1='"tecnico,funzionale"',
                    allow_blank=False, showDropDown=False,
                    error="Valore non valido: usa 'tecnico' o 'funzionale'",
                    errorTitle="Origine bug",
                    prompt="Origine del bug: tecnico (test automatici team tech) o funzionale (playbook funzionale)",
                    promptTitle="Origine")
dv.add("J2:J1000")
ws.add_data_validation(dv)

# --- Sheet Istruzioni ---
ws_i = wb["Istruzioni"]

# Sostituisci lo schema riga e aggiungi note
for r in range(1, ws_i.max_row + 1):
    v = ws_i.cell(row=r, column=1).value
    if v is None:
        continue
    if "id, fase, sezione, utente, titolo" in v and "origine" not in v:
        ws_i.cell(row=r, column=1,
                  value=("  id, fase, sezione, utente, titolo, descrizione, screenshot, "
                         "riferimento, tipo, origine, stato_originale, data, note_dev, "
                         "note_funzionale"))
    elif "Fase 11" in v:
        ws_i.cell(row=r, column=1, value=v.replace("Fase 11", "Fase 2c (test e chiusura)"))
    elif "stato_originale default Aperto" in v:
        # aggiungi bullet origine appena dopo 'tipo deve essere...'
        pass  # gestito sotto

# Aggiungi sezione dedicata a 'origine' come ultime righe esplicative del foglio
add_rows = [
    "",
    "Colonna origine (NUOVA v2):",
    "  - tecnico    -> bug rilevato dai test tecnici automatici (unit/integration/perf/security) lanciati dal team tech in Fase 2c ondata (a).",
    "  - funzionale -> bug rilevato dal team funzionale in autonomia eseguendo il playbook test (md/xlsx) in Fase 2c ondata (b).",
    "  - sdlc-debug raggruppa i bug per origine in BUG_REPORT.md e applica counter separati per la chiusura plan.",
    "  - Per file legacy senza colonna origine (template v1), sdlc-debug applica default origine=tecnico.",
]
start_row = ws_i.max_row + 1
for offset, text in enumerate(add_rows):
    ws_i.cell(row=start_row + offset, column=1, value=text if text else None)

# --- Save ---
DST.parent.mkdir(parents=True, exist_ok=True)
wb.save(DST)
print(f"Saved: {DST}")
print(f"Bugs sheet now has {wb['Bugs'].max_column} columns and {wb['Bugs'].max_row} rows.")
