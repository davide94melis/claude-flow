#!/usr/bin/env python3
"""generate-progress-xlsx.py — deterministic manifest-driven renderer for the SDLC
progress Excel report (sdlc-progress-report skill; SDLC feedback round #6/#3/#4).

Renders ANY progress manifest (official default or per-project custom) into PROGRESS.xlsx
from a plan's TASKS.md + PROGRESS.md (+ optional PLAN.md header for deadline/cadence).

Centralized model: the caller (skill) fetches TASKS/PROGRESS from origin/main and passes
their paths here; this script never touches git.

Subcommands:
  render   parse inputs, render sheets per manifest, save xlsx.
  analyze  read a TL-provided .xlsx and emit a draft manifest (best-effort field bindings).

Anti-zeroing guard (#4): render refuses (exit 3) to overwrite an existing xlsx whose Task
sheet already has non-null progress with an all-zero/empty dataset, unless --allow-zero.
"""
import argparse
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta

DEFAULT_MANIFEST = {
    "version": 1,
    "name": "official-default",
    "sheets": [
        {
            "name": "Task",
            "type": "table",
            "columns": [
                {"header": "ID", "width": 10, "field": "id"},
                {"header": "Stream", "width": 18, "field": "stream"},
                {"header": "Attività", "width": 30, "field": "activity"},
                {"header": "Descrizione", "width": 60, "field": "description", "wrap": True},
                {"header": "Owner", "width": 18, "field": "owner"},
                {"header": "Area", "width": 8, "field": "area"},
                {"header": "Priorità", "width": 10, "field": "priority"},
                {"header": "Wave", "width": 10, "field": "wave"},
                {"header": "Dipendenze", "width": 15, "field": "dependencies"},
                {"header": "Effort", "width": 10, "field": "effort"},
                {"header": "Branch", "width": 25, "field": "branch"},
                {"header": "Progresso", "width": 12, "field": "progress", "format": "progress_bands"},
                {"header": "Stato", "width": 15, "field": "status", "format": "status_colors"},
                {"header": "Note", "width": 40, "field": "note", "wrap": True, "preserve_manual": True},
            ],
        },
        {"name": "Per Sviluppatore", "type": "by_developer"},
        {"name": "Riepilogo", "type": "summary"},
    ],
}

FIELD_ALIASES = {
    "id": "id", "stream": "stream",
    "attivita": "activity", "attivita'": "activity", "attivita’": "activity", "activity": "activity",
    "descrizione": "description", "description": "description",
    "owner": "owner", "sviluppatore": "owner", "developer": "owner",
    "area": "area", "priorita": "priority", "priorita'": "priority", "priority": "priority",
    "wave": "wave", "dipendenze": "dependencies", "dependencies": "dependencies",
    "effort": "effort", "stima": "effort",
    "branch": "branch", "progresso": "progress", "progress": "progress", "%": "progress",
    "stato": "status", "status": "status", "note": "note", "nota": "note", "notes": "note",
}

DONE = "Completata"
STATUS_ACTIVE = {"In corso"}
STATUS_BLOCKED = {"Bloccata"}
STATUS_INACTIVE = {"Annullata", "Sospesa"}
STATUS_CANON = {
    "completata": "Completata", "in corso": "In corso", "bloccata": "Bloccata",
    "da iniziare": "Da iniziare", "annullata": "Annullata", "sospesa": "Sospesa",
}


def canon_status(s):
    """Map a free-form status to a canonical label (case/accent tolerant) so downstream
    exact comparisons don't drop a task from cadence/summary/per-dev on a stray-case value."""
    return STATUS_CANON.get(norm(s), (s or "").strip() or "Da iniziare")


def norm(s):
    s = (s or "").strip().lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s.strip()


def parse_md_tables(text):
    """Return list of tables; each table = list of row-dicts keyed by RAW header text."""
    tables, headers, sep_seen = [], None, False
    cur = []
    for line in (text or "").split("\n"):
        st = line.strip()
        if st.startswith("|") and st.endswith("|"):
            cells = [c.strip() for c in st.strip("|").split("|")]
            if headers is None:
                headers, sep_seen, cur = cells, False, []
                continue
            if not sep_seen and set("".join(cells)) <= set("-: "):
                sep_seen = True
                continue
            row = {headers[i]: (cells[i] if i < len(cells) else "") for i in range(len(headers))}
            cur.append(row)
        else:
            if headers is not None and cur:
                tables.append((headers, cur))
            headers, sep_seen, cur = None, False, []
    if headers is not None and cur:
        tables.append((headers, cur))
    return tables


def _rowmap(headers):
    return {norm(h): h for h in headers}


def _pick(row, hmap, *keys):
    for k in keys:
        if k in hmap:
            return row.get(hmap[k], "")
    return ""


def parse_tasks(tasks_text):
    """Parse TASKS.md: the table whose headers contain an 'ID' column."""
    out = {}
    for headers, rows in parse_md_tables(tasks_text):
        hmap = _rowmap(headers)
        if "id" not in hmap:
            continue
        for row in rows:
            tid = _pick(row, hmap, "id").strip()
            if not re.match(r"^T-[A-Z0-9-]+$", tid):
                continue
            out[tid] = {
                "id": tid,
                "stream": _pick(row, hmap, "stream"),
                "activity": _pick(row, hmap, "attivita", "activity"),
                "description": _pick(row, hmap, "descrizione", "description"),
                "owner": _pick(row, hmap, "owner", "sviluppatore"),
                "area": _pick(row, hmap, "area"),
                "priority": _pick(row, hmap, "priorita", "priority"),
                "wave": _pick(row, hmap, "wave"),
                "dependencies": _pick(row, hmap, "dipendenze", "dependencies"),
                "effort": _pick(row, hmap, "effort", "stima"),
                "branch": _pick(row, hmap, "branch"),
                "progress": "0", "status": "Da iniziare", "note": "",
            }
    return out


def parse_progress(progress_text, tasks):
    """Merge PROGRESS.md status/progress/branch/note into tasks (by ID)."""
    for headers, rows in parse_md_tables(progress_text):
        hmap = _rowmap(headers)
        if "id" not in hmap:
            continue
        for row in rows:
            tid = _pick(row, hmap, "id").strip()
            if tid not in tasks:
                continue
            prog = _pick(row, hmap, "progresso", "progress")
            m = re.search(r"(\d+)", str(prog))
            if m:
                tasks[tid]["progress"] = m.group(1)
            status = _pick(row, hmap, "stato", "status").strip()
            if status:
                tasks[tid]["status"] = canon_status(status)
            br = _pick(row, hmap, "branch").strip()
            if br:
                tasks[tid]["branch"] = br
            note = _pick(row, hmap, "note", "nota").strip()
            if note:
                tasks[tid]["note"] = note
    return tasks


def parse_plan_dates(plan_text):
    start = deadline = None
    for label, key in (("data inizio ufficiale", "start"), ("deadline", "deadline")):
        m = re.search(label + r"\s*[:：]?\s*`?(\d{4}-\d{2}-\d{2})`?", plan_text or "", re.IGNORECASE)
        if m:
            try:
                d = datetime.strptime(m.group(1), "%Y-%m-%d").date()
            except ValueError:
                d = None
            if key == "start":
                start = d
            else:
                deadline = d
    return start, deadline


def workdays(d1, d2):
    if not d1 or not d2 or d2 < d1:
        return 0
    n, cur = 0, d1
    while cur <= d2:
        if cur.weekday() < 5:
            n += 1
        cur += timedelta(days=1)
    return n


def to_float(x):
    m = re.search(r"[-+]?\d*\.?\d+", str(x).replace(",", "."))
    return float(m.group(0)) if m else 0.0


def to_int(x):
    m = re.search(r"\d+", str(x))
    return int(m.group(0)) if m else 0


def load_manifest(spec):
    if not spec or spec == "official":
        return DEFAULT_MANIFEST
    with open(spec, encoding="utf-8") as fh:
        return json.load(fh)


# ---------- rendering ----------
def render(args):
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return 2

    tasks_text = open(args.tasks, encoding="utf-8").read() if args.tasks and os.path.exists(args.tasks) else ""
    progress_text = open(args.progress, encoding="utf-8").read() if args.progress and os.path.exists(args.progress) else ""
    plan_text = open(args.plan, encoding="utf-8").read() if args.plan and os.path.exists(args.plan) else ""

    tasks = parse_tasks(tasks_text)
    parse_progress(progress_text, tasks)
    records = [tasks[k] for k in sorted(tasks)]

    manifest = load_manifest(args.manifest)

    # Locate the task-table sheet + map fields -> header text, so the anti-zeroing guard
    # and note-preservation work for CUSTOM templates too (not just the official layout).
    table_sheet = next((s for s in manifest.get("sheets", []) if s.get("type", "table") == "table"), None)
    tbl_cols = table_sheet.get("columns", []) if table_sheet else []
    tbl_name = table_sheet["name"] if table_sheet else "Task"
    field_header = {c["field"]: c.get("header", "") for c in tbl_cols if c.get("field")}
    note_preserve = table_sheet is None or any(
        c.get("field") == "note" and c.get("preserve_manual") for c in tbl_cols
    )
    id_hdr = norm(field_header.get("id", "id"))
    prog_hdr = norm(field_header.get("progress", "progresso"))
    note_hdr = norm(field_header.get("note", "note"))

    non_null = any(to_int(r["progress"]) > 0 or r["status"] == DONE for r in records)
    existing_non_null = False
    prev_notes = {}
    if os.path.exists(args.out):
        try:
            wbp = load_workbook(args.out)
            if tbl_name in wbp.sheetnames:
                ws0 = wbp[tbl_name]
                hdr = [c.value for c in ws0[1]]
                idx = {norm(str(h)): i for i, h in enumerate(hdr) if h}
                pi, ni, ii = idx.get(prog_hdr), idx.get(note_hdr), idx.get(id_hdr)
                for row in ws0.iter_rows(min_row=2, values_only=True):
                    if ii is not None and pi is not None and to_int(row[pi]) > 0:
                        existing_non_null = True
                    if ii is not None and ni is not None and row[ii] and row[ni]:
                        prev_notes[str(row[ii]).strip()] = str(row[ni])
        except Exception:
            pass

    if (not records or not non_null) and existing_non_null and not args.allow_zero:
        print(
            "STOP (anti-zeroing guard #4): dataset vuoto o tutto-0% ma esiste già un PROGRESS.xlsx "
            "con progresso non-nullo. Probabile sync fallita o PROGRESS.md non trovato/parsato. "
            "Non sovrascrivo. Usa --allow-zero solo per un piano genuinamente non iniziato.",
            file=sys.stderr,
        )
        return 3

    # carry over manual notes where new note is empty (only if the manifest opts in via preserve_manual)
    if note_preserve:
        for r in records:
            if not r["note"] and r["id"] in prev_notes:
                r["note"] = prev_notes[r["id"]]

    today = datetime.strptime(args.today, "%Y-%m-%d").date() if args.today else date.today()

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="404040")
    bands = {"zero": "F4CCCC", "low": "FCE5CD", "mid": "FFF2CC", "full": "D9EAD3"}
    status_fill = {
        DONE: "D9EAD3", "In corso": "CFE2F3", "Bloccata": "F4CCCC",
        "Annullata": "D9D9D9", "Sospesa": "D9D9D9",
    }

    for sheet in manifest.get("sheets", []):
        stype = sheet.get("type", "table")
        ws = wb.create_sheet(sheet["name"])
        if stype == "table":
            cols = sheet.get("columns", [])
            for ci, col in enumerate(cols, 1):
                c = ws.cell(1, ci, col.get("header", ""))
                c.font, c.fill = header_font, header_fill
                ws.column_dimensions[get_column_letter(ci)].width = col.get("width", 15)
            for ri, r in enumerate(records, 2):
                for ci, col in enumerate(cols, 1):
                    field = col.get("field")
                    val = r.get(field, "") if field else ""
                    if field == "progress":
                        val = f"{to_int(val)}%"
                    cell = ws.cell(ri, ci, val)
                    if col.get("wrap"):
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                    if col.get("format") == "progress_bands" and field:
                        p = to_int(r.get(field, 0))
                        key = "zero" if p == 0 else "low" if p < 50 else "mid" if p < 100 else "full"
                        cell.fill = PatternFill("solid", fgColor=bands[key])
                    if col.get("format") == "status_colors" and field:
                        f = status_fill.get(str(r.get(field, "")).strip())
                        if f:
                            cell.fill = PatternFill("solid", fgColor=f)
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(len(records) + 1, 1)}"
        elif stype == "by_developer":
            _render_by_developer(ws, records, header_font, header_fill)
        elif stype == "summary":
            _render_summary(ws, records, plan_text, today, manifest, header_font)
    wb.save(args.out)
    print(f"Written {args.out} ({len(records)} tasks, manifest='{manifest.get('name')}')")
    return 0


def _render_by_developer(ws, records, header_font, header_fill):
    headers = ["Sviluppatore", "Ruolo", "Task totali", "Completate", "In corso",
               "Da iniziare", "Bloccate", "Progresso medio", "Effort totale", "Effort completato"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(1, ci, h)
        c.font, c.fill = header_font, header_fill
    by = {}
    for r in records:
        by.setdefault(r["owner"] or "—", []).append(r)
    ri = 2
    for owner, rs in sorted(by.items()):
        eff = sum(to_float(x["effort"]) for x in rs)
        effc = sum(to_float(x["effort"]) for x in rs if x["status"] == DONE)
        avg = round(sum(to_int(x["progress"]) for x in rs) / max(len(rs), 1))
        vals = [owner, rs[0]["area"], len(rs),
                sum(1 for x in rs if x["status"] == DONE),
                sum(1 for x in rs if x["status"] in STATUS_ACTIVE),
                sum(1 for x in rs if x["status"] == "Da iniziare"),
                sum(1 for x in rs if x["status"] in STATUS_BLOCKED),
                f"{avg}%", round(eff, 1), round(effc, 1)]
        for ci, v in enumerate(vals, 1):
            ws.cell(ri, ci, v)
        ri += 1


def _render_summary(ws, records, plan_text, today, manifest, header_font):
    total = len(records)
    done = sum(1 for r in records if r["status"] == DONE)
    active = sum(1 for r in records if r["status"] in STATUS_ACTIVE)
    todo = sum(1 for r in records if r["status"] == "Da iniziare")
    blocked = sum(1 for r in records if r["status"] in STATUS_BLOCKED)
    inactive = sum(1 for r in records if r["status"] in STATUS_INACTIVE)
    overall = round(sum(to_int(r["progress"]) for r in records) / max(total, 1))
    eff_tot = sum(to_float(r["effort"]) for r in records)
    eff_done = sum(to_float(r["effort"]) for r in records if r["status"] == DONE)  # done-only: reconciles with Per Sviluppatore
    eff_progress = sum(to_float(r["effort"]) * to_int(r["progress"]) / 100.0 for r in records)  # progress-weighted: cadence measure

    def pct(n):
        return f"{round(100 * n / max(total, 1))}%"

    lines = [
        ("Data generazione", today.isoformat()),
        ("", ""),
        ("STATO COMPLESSIVO", ""),
        ("Task totali", total),
        ("Completate", f"{done} ({pct(done)})"),
        ("In corso", f"{active} ({pct(active)})"),
        ("Da iniziare", f"{todo} ({pct(todo)})"),
        ("Bloccate", f"{blocked} ({pct(blocked)})"),
        ("Annullate/Sospese", f"{inactive} ({pct(inactive)})"),
        ("Progresso complessivo", f"{overall}%"),
        ("", ""),
        ("EFFORT", ""),
        ("Effort totale stimato (gg)", round(eff_tot, 1)),
        ("Effort completato (gg)", round(eff_done, 1)),
        ("Effort rimanente (gg)", round(eff_tot - eff_done, 1)),
    ]

    start, deadline = parse_plan_dates(plan_text)
    if start and deadline:
        n_tot = workdays(start, deadline)
        n_elapsed = workdays(start, min(today, deadline))
        n_left = max(workdays(today, deadline) - (1 if today.weekday() < 5 else 0), 0)
        req_cadence = eff_tot / n_tot if n_tot else 0
        expected = eff_tot * n_elapsed / n_tot if n_tot else 0
        delta_eff = eff_progress - expected
        status = "ANTICIPO" if delta_eff >= 0 else "RITARDO"
        velocity = eff_progress / n_elapsed if n_elapsed else 0
        proj_days = (eff_tot - eff_progress) / velocity if velocity > 0 else None
        if proj_days is not None:
            proj_end, added = today, 0
            while added < round(proj_days):
                proj_end += timedelta(days=1)
                if proj_end.weekday() < 5:
                    added += 1
            proj_str = proj_end.isoformat()
            slip = workdays(deadline, proj_end) - 1 if proj_end > deadline else -(workdays(proj_end, deadline) - 1)
        else:
            proj_str, slip = "n/d (velocità 0)", "n/d"
        lines += [
            ("", ""),
            ("CADENZA VERSO LA DEADLINE (baseline lineare, weekend esclusi)", ""),
            ("Data inizio ufficiale", start.isoformat()),
            ("Deadline", deadline.isoformat()),
            ("Giorni lavorativi totali", n_tot),
            ("Giorni lavorativi trascorsi", n_elapsed),
            ("Giorni lavorativi rimanenti", n_left),
            ("Cadenza richiesta (gg-effort/giorno)", round(req_cadence, 2)),
            ("Cadenza richiesta (task/giorno)", round(total / n_tot, 2) if n_tot else 0),
            ("Effort atteso a oggi (gg)", round(expected, 1)),
            ("Effort effettivo a oggi (gg, progress-weighted)", round(eff_progress, 1)),
            (f"{status} (delta gg-effort)", round(delta_eff, 1)),
            ("Data-fine proiettata", proj_str),
            ("Scarto vs deadline (gg lav.; +=ritardo, -=anticipo)", slip),
        ]
    else:
        lines += [("", ""), ("CADENZA", "date non fornite nell'header del PLAN — blocco ANTICIPO/RITARDO omesso")]

    for ri, (k, v) in enumerate(lines, 1):
        a = ws.cell(ri, 1, k)
        if k and not v:
            a.font = header_font
        ws.cell(ri, 2, v)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 40


# ---------- analyze ----------
def analyze(args):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        return 2
    wb = load_workbook(args.xlsx, read_only=True)
    sheets = []
    for i, ws in enumerate(wb.worksheets):
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        columns = []
        for j, h in enumerate(header_row):
            if h is None:
                continue
            field = FIELD_ALIASES.get(norm(str(h)))
            col = {"header": str(h), "width": 18}
            col["field"] = field  # None if unmapped -> blank column (documented)
            columns.append(col)
        sheets.append({"name": ws.title, "type": "table" if i == 0 else "custom", "columns": columns})
    manifest = {"version": 1, "name": "custom-draft", "source_xlsx": os.path.basename(args.xlsx), "sheets": sheets}
    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(manifest, fh, ensure_ascii=False, indent=2)
    print(f"Draft manifest written to {args.out}. Review field bindings (null = blank column).")
    return 0


def main(argv=None):
    p = argparse.ArgumentParser(description="SDLC progress-report renderer (manifest-driven).")
    sub = p.add_subparsers(dest="cmd")
    r = sub.add_parser("render")
    r.add_argument("--tasks", required=True)
    r.add_argument("--progress")
    r.add_argument("--plan")
    r.add_argument("--manifest", default="official")
    r.add_argument("--out", required=True)
    r.add_argument("--today")
    r.add_argument("--allow-zero", action="store_true")
    a = sub.add_parser("analyze")
    a.add_argument("--xlsx", required=True)
    a.add_argument("--out", required=True)
    args = p.parse_args(argv)
    if args.cmd == "analyze":
        return analyze(args)
    if args.cmd == "render":
        return render(args)
    p.print_help()
    return 2


if __name__ == "__main__":
    sys.exit(main())
